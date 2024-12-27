import os
import requests
import pandas as pd
from openpyxl import load_workbook

def get_fgt_credentials(input_file):
    try:
        workbook = load_workbook(input_file)
        if 'fgt_ip' not in workbook.sheetnames:
            raise ValueError("The input file does not contain a 'fgt_ip' tab.")
        if 'statistic' not in workbook.sheetnames:
            raise ValueError("The input file does not contain a 'statistic' tab.")
        
        fgt_ip_data = pd.read_excel(input_file, sheet_name='fgt_ip')
        if 'FortiGate_IP' not in fgt_ip_data.columns or 'Bearer_Token' not in fgt_ip_data.columns:
            raise ValueError("The 'fgt_ip' tab must contain 'FortiGate_IP' and 'Bearer_Token' columns.")
        
        fortigate_credentials = list(fgt_ip_data[['FortiGate_IP', 'Bearer_Token']].dropna().itertuples(index=False, name=None))

        statistic_data = pd.DataFrame({
            #'FortiGate_IP': [ip for ip, _ in fortigate_credentials],
            'FortiGate_IP': [None] * len(fortigate_credentials),
        })

        with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            statistic_data.to_excel(writer, sheet_name='statistic', index=False)
        return fortigate_credentials
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

def get_current_usage(data):
    current_usages = {}
    if 'results' in data:
        for key, value in data['results'].items():
            if isinstance(value, dict) and 'current_usage' in value:
                current_usages[key] = value['current_usage']
    return current_usages

def get_fgt_performance(fgt_ip, token):
    url = f"https://{fgt_ip}/api/v2/monitor/system/performance/status"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    try:
        response = requests.get(url, headers=headers, verify=False)
        if response.status_code == 200:
            #print(f"Performance data for {fgt_ip}:")
            return response.json()
        else:
            return {"error": f"Failed to fetch data. Status Code: {response.status_code}"}
    except requests.exceptions.RequestException as e:
        print(f"An error occurred while fetching data from {fgt_ip}: {e}")
        return {"error": str(e)}
    
def get_fgt_resources(fgt_ip, token):
    url = f"https://{fgt_ip}/api/v2/monitor/system/global-resources"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }    
    try:
        response = requests.get(url, headers=headers, verify=False)
        if response.status_code == 200:
            #print(f"Resources data for {fgt_ip}:")
            return response.json()
        else:
            return {"error": f"Failed to fetch data. Status Code: {response.status_code}"}
    except requests.exceptions.RequestException as e:
        print(f"An error occurred while fetching data from {fgt_ip}: {e}")
        return {"error": str(e)}
    
def update_statistic_tab(file_path, fgt_ip, current_usages):
    try:
        workbook = load_workbook(file_path)
        if 'statistic' not in workbook.sheetnames:
            raise ValueError("The input file does not contain a 'statistic' tab.")
        
        statistic_data = pd.read_excel(file_path, sheet_name='statistic')
        if 'FortiGate_IP' not in statistic_data.columns:
            raise ValueError("The 'statistic' tab must contain a 'FortiGate_IP' column.")

        # Ensure all keys in current_usages are present as headers
        for usage_key in current_usages.keys():
            if usage_key not in statistic_data.columns:
                statistic_data[usage_key] = None  # Add missing columns

        '''
        # Update the row corresponding to fgt_ip or add it if not present
        if fgt_ip in statistic_data['FortiGate_IP'].values:
            # Update existing row
            row_idx = statistic_data.loc[statistic_data['FortiGate_IP'] == fgt_ip].index[0]
            for key, value in current_usages.items():
                statistic_data.at[row_idx, key] = value
        else:
        '''
        # Add a new row for the IP
        new_row = {'FortiGate_IP': fgt_ip, **current_usages}
        statistic_data = pd.concat([statistic_data, pd.DataFrame([new_row])], ignore_index=True)

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            statistic_data.to_excel(writer, sheet_name='statistic', index=False)
        print(f"Updated 'statistic' tab for {fgt_ip} in {file_path}.")
        
    except Exception as e:
        print(f"An error occurred: {e}")
    
os.system('clear')
filename = 'fgt_stat.xlsx'
current_directory = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(current_directory, filename)

credentials = get_fgt_credentials(file_path)
#print(f"FortiGate Credentials: {credentials}")

for fgt_ip, token in credentials:
    data = get_fgt_resources(fgt_ip, token)
    if "error" not in data:
       update_statistic_tab(file_path, fgt_ip, get_current_usage(data))
       #print(current_usages) 
    else:
       print(f"Error for {fgt_ip}:", data["error"])
    